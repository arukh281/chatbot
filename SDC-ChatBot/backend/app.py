from flask import Flask, request, jsonify
from flask_cors import CORS  # Import CORS
from response import chatbot_response  # Import your chatbot function
import openpyxl

app = Flask(__name__)
CORS(app)  # Enable CORS

# Load or create the Excel workbook
workbook_path = "feedback.xlsx"
try:
    workbook = openpyxl.load_workbook(workbook_path)
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    workbook.save(workbook_path)

# Ensure required sheets exist
if "Doubtful" not in workbook.sheetnames:
    workbook.create_sheet("Doubtful")
if "Not Helpful" not in workbook.sheetnames:
    workbook.create_sheet("Not Helpful")
workbook.save(workbook_path)

def save_to_worksheet(sheet_name, question, answer):
    sheet = workbook[sheet_name]
    sheet.append([question, answer])
    workbook.save(workbook_path)

@app.route('/chat', methods=['POST'])
def chat():
    data = request.get_json()
    user_message = data.get('message', '')
    confidence, response = chatbot_response(user_message)  # Fix the order here

    return jsonify({
        'response': response,
        'confidence': confidence
    })

@app.route('/feedback', methods=['POST'])
def feedback():
    data = request.get_json()
    question = data.get('question', '')
    answer = data.get('answer', '')
    feedback_type = data.get('feedback_type', '')  # "doubtful" or "not_helpful"

    print(f"Received feedback: question='{question}', answer='{answer}', feedback_type='{feedback_type}'")  # Debugging

    if feedback_type == "doubtful":
        save_to_worksheet("Doubtful", question, answer)
    elif feedback_type == "not_helpful":
        save_to_worksheet("Not Helpful", question, answer)

    return jsonify({"message": "Feedback saved successfully"})

@app.route('/', methods=['GET'])
def home():
    return "University FAQ Chatbot API is running. Send POST requests to /chat"

if __name__ == '__main__':
    app.run(debug=True, port=5001)
